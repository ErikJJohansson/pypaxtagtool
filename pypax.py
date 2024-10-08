from pycomm3 import LogixDriver
from sys import argv
import openpyxl
from tqdm import trange, tqdm
from itertools import product
import argparse

'''

    argv 1 path to excel file (REQUIRED to be backwards compatible
    argv 2 PLC path if we want to be backwards compatible without changing spreadsheet
    argv 3 Check for -r or -w for read/write

'''
# These are constants based on the PlantPAX spreadsheets containing AOI data
START_ROW = 10
START_COL = 5
NAME_COL = 3
TOP_TAG_ROW = 7
BOTTOM_TAG_ROW = 8

# These are constants based on the PlantPAX setup sheet at the beginning of the workbook

NUM_INSTANCES_COL = 4 # in setup sheet
NUM_SUBTAGS_COL = 5  # in setup sheet

def get_aoi_tag_instances(plc, tag_type):
    """
    function returns list of tag names matching struct type
    """
    #return tag_list

    tag_list = []

    for tag, _def in plc.tags.items():
        if _def['data_type_name'] == tag_type and not(_def['alias']):
            if _def['dim'] > 0:
                tag_list = tag_list + get_dim_list(tag,_def['dimensions'])
            else:
                tag_list.append(tag)

    return tag_list

def get_aoi_list(excel_book):
    aoi_list = []

    # PlantPAX AOI's have an _ for second character
    for sheet in excel_book.sheetnames:
        if sheet[1] == '_' or (sheet[0] == 'S' and sheet[1] == 'I' and sheet[2] == 'F'):
            aoi_list.append(sheet)

    return aoi_list

def get_subtag_list(sheet):
    '''
    function gets all subtags in a given sheet, returns a list of subtags
    '''
       
    sub_tag_list = [] 
    i = START_COL
    sub_tag = get_subtag(sheet,i)

    while sub_tag != 'NoneNone':            
        
        sub_tag_list.append(sub_tag)
        
        #update iterator
        i+=1
        sub_tag = get_subtag(sheet,i)

    return sub_tag_list

def get_subtag(sheet, column):
    '''
    function gets subtag based on column
    '''

    sub_tag = str(sheet.cell(TOP_TAG_ROW,column).value) + str(sheet.cell(BOTTOM_TAG_ROW,column).value)

    return sub_tag

def search_value_in_col(sheet, search_string, col_idx=1):
    '''
    search a column for the specific string, return row on match
    '''
    for row in range(1, sheet.max_row + 1):
        if sheet.cell(row,col_idx).value == search_string:
            return row
    
    return None

def get_aoi_setup(sheet):
    '''
    Finds the number of AOIs in a current sheet
    '''

    i = 0

    while True:                   

        # read row
        cell_value = sheet.cell(START_ROW+i,NAME_COL).value

        # conditions to see if it's a legit value
        tag_exists = (cell_value != None) or (cell_value == '')

        # NoneNone means the cell is blank, so we assume we are done
        if tag_exists:
            i += 1
        else:
            break

    num_aoi_tags = i

    return num_aoi_tags

def set_num_instances(sheet, aoi_name, num):
    '''
    Function updates the num instances in the Setup page of spreadsheet
    '''

    # 7 is col H "Worksheet tab name"
    aoi_row = search_value_in_col(sheet, aoi_name, 8)
    
    if aoi_row != None:
        sheet.cell(aoi_row,4).value = num

def get_dim_list(base_tag, dim_list):
    '''
    function takes a list which has the array size and turns it into a list with all iterations
    '''
    # remove 0's
    filtered_list = list(filter(lambda num: num != 0, dim_list))

    temp = []

    for indices in product(*[range(dim) for dim in filtered_list]):
        temp.append(base_tag + ''.join(f'[{i}]' for i in indices))

    return temp

def make_tag_list(base_tag,sub_tags):
    '''
    returns the full tag path of a given base tag and sub tags
    '''
    # concatenate base tag
    read_list = [base_tag + s for s in sub_tags]

    return read_list

def read_plc_row(plc, tag_list):
    '''
    reads data from plc, returns list of tuples (tag_name, tag_value)
    '''
    
    result = tag_data = plc.read(*tag_list)

    # tuple of tag name, data
    tag_data_formatted = []

    # hardcoded but it works
    for s in tag_data:
        if s[2] == 'BOOL':
            data = int(s[1])
        elif s[2] == 'REAL':
            if 'e' in str(s[1]):
                data = float(format(s[1],'.6e'))
            elif s[1].is_integer():
                data = int(s[1])
            else:
                data = float(format(s[1], '.6f'))                  
        else:
            data = s[1]

        tag_data_formatted.append((s[0],data))

    return tag_data_formatted, result

def write_plc_row(plc, tag_data):
    '''
    writes tag data to tags in plc
    '''
    result = plc.write(*tag_data)

    # added to always format the result as a list
    if len(tag_data) == 1:
        return [result]
    else:
        return result

def write_sheet_row(sheet,row,base_tag,tag_data):
    '''
    writes tag data to a row in spreadsheet
    '''
    # write name
    sheet.cell(row,NAME_COL).value = base_tag

    # write data    
    for i in range(len(tag_data)):
            
        sheet.cell(row,START_COL+i).value = tag_data[i][1]

def read_data_sheet_row(sheet,row,sub_tags):
    '''
    reads tag name and data from list
    '''
    base_tag = sheet.cell(row,NAME_COL).value

    tag_data = []

    # loop through subtags, get data
    for i in range(len(sub_tags)):

        if sheet.cell(row,START_COL+i).value == None:
            cell_value = (base_tag + sub_tags[i],'')
        else:
            cell_value = (base_tag + sub_tags[i],sheet.cell(row,START_COL+i).value)

        tag_data.append(cell_value)

    return base_tag,tag_data

def failed_tag_formatter(tags, write_mode):
    '''
    Function takes list of failed tags and returns string
    '''
    if write_mode:
        return '****WARNING****\nCANNOT WRITE TO\n' + ", ".join(tags) + '\n***************'
    else:
        return '****WARNING****\nCANNOT READ FROM\n' + ", ".join(tags) + '\n***************'       

def get_failed_tags(tag_list, result_list):
    '''
    function takes a list of booleans and only returns tags that had issues
    '''
    failed_indexes = []
    failed_tags = []

    failed_indexes = [i for i, val in enumerate(result_list) if not val]
    failed_tags = failed_tags + [tag_list[i] for i in failed_indexes]

    return failed_tags

def main():

    # default filename of template file included in the repo
    default_excelfile = 'ProcessLibraryOnlineConfigTool.xlsm'

    # Parse arguments
   
    parser = argparse.ArgumentParser(
        description='Python-based PlantPAX tag configuration tool.',
        epilog='This tool works on both Windows and Mac.')
    
    # Add command-line arguments
    parser.add_argument('commpath', help='Path to PLC')
    subparsers = parser.add_subparsers(dest='mode',help='Select read/write mode')

    # parsing read commands, filename is optional and will default to default_excelfile value
    read_parser = subparsers.add_parser('read', help='Read tags from PLC into spreadsheet')
    read_parser.add_argument('excelfile', nargs='?', default=default_excelfile,help='Path to excel file')

    # parsing write commands, excelfile is required
    write_parser = subparsers.add_parser('write', help='Write data from spreadsheet into PLC tags')
    write_parser.add_argument('excelfile',help='Path to excel file')
                                       
    args = parser.parse_args()

    # Access the parsed arguments
    commpath = args.commpath
    excelfile = args.excelfile
    mode = args.mode

    # open connection to PLC

    plc = LogixDriver(commpath, init_tags=True,init_program_tags=True)

    print('Connecting to PLC.')
    try:
        plc.open()
        plc_name = plc.get_plc_name()

        print('Connected to ' + plc_name + ' PLC at ' + commpath)
    except:
        print('Unable to connect to PLC at ' + commpath)
        exit()

    # open excel file

    # filename check
    if mode == 'write' and excelfile.find(plc_name) == -1:
        print("Filename mismatch. The file '" + excelfile + "' does not contain '" + plc_name + "'.")
        exit()

    print('Opening ' + excelfile)
    try:
        book = openpyxl.load_workbook(excelfile,keep_vba=False,keep_links=True)

    except:
        print('Unable to open excel file ' + excelfile)
        plc.close()
        exit()
    
    print('Opened file named ' + excelfile)

    # get list of AOI sheet names
    aoi_sheet_names = get_aoi_list(book)

    # should be the first sheet in the workbook
    setup_sheet = book["Setup"]

    # read from PLC
    if mode == 'read':
        print('Reading tags from ' + plc_name + ' PLC.')
        
        for aoi in aoi_sheet_names:
            # get setup info from PLC tags, write to spreadsheet
            base_tags = get_aoi_tag_instances(plc,aoi)
            num_instances = len(base_tags)
            set_num_instances(setup_sheet,aoi,num_instances)

            if num_instances > 0:

                # get subtag list for given AOI
                sub_tags = get_subtag_list(book[aoi])

                failed_read_tags = []

                # read rows, write to spreadsheet
                for i in tqdm(range(num_instances),"Reading instances of " + aoi):
                    tag_list = make_tag_list(base_tags[i],sub_tags)

                    # data for one tag and all sub tags
                    tag_data, read_result = read_plc_row(plc,tag_list)

                    # add to failed tags list if we can't find the tag
                    if not all(read_result):
                        failed_read_tags = failed_read_tags + get_failed_tags(tag_list,read_result)

                    write_sheet_row(book[aoi],START_ROW+i,base_tags[i],tag_data)

                # print to command line if we couldn't read any tags
                if failed_read_tags:
                    print(failed_tag_formatter(failed_read_tags,False))
            else:
                print("No instances of " + aoi + " found in " + plc_name + " PLC.")

        # add plc name to file and save to new file
        outfile = plc_name + '_ConfigTags.' + 'xlsx'
        print('Finished reading from ' + plc_name + ' PLC.')
        print('Saving to file ' + outfile)
        book.save(outfile)
        print('file saved to ' + outfile)

    # Write to PLC
    elif mode == 'write':
        print('Writing tags to ' + plc_name + ' PLC.')
        
        for aoi in aoi_sheet_names:

            # get aoi info from sheet and plc
            num_instances_in_sheet = get_aoi_setup(book[aoi])
            base_tags = get_aoi_tag_instances(plc,aoi)

            # Check to make sure there are instances in sheet
            if num_instances_in_sheet > 0:

                # get subtags of datatype
                sub_tags = get_subtag_list(book[aoi])

                # reset lists
                tag_data_differences = []       
                failed_read_tags = []
                failed_write_tags   = []

                # read AOI data from each row in spreadsheet, add differences to tag_data_differences
                for i in tqdm(range(num_instances_in_sheet),"Comparing instances of " + aoi):

                    # data for one tag and all sub tags from sheet
                    base_tag, tag_data_sheet = read_data_sheet_row(book[aoi],START_ROW+i,sub_tags)

                    # data for one tag and all sub tags from plc
                    tag_list = make_tag_list(base_tag,sub_tags)
                    tag_data_plc, read_result = read_plc_row(plc,tag_list)

                    # add to failed tags list if we can't find the tag
                    if not all(read_result):
                        failed_read_tags = failed_read_tags + [base_tag] #get_failed_tags(tag_list,read_result)
                    
                    else:
                        # compare PLC row to spreadsheet row, add differences to list if any for that tag
                        row_differences = list(set(tag_data_sheet).difference(set(tag_data_plc)))
                        if row_differences:
                            tag_data_differences += row_differences

                # print to command line if we couldn't read any tags for that instance
                if failed_read_tags:
                    print(failed_tag_formatter(failed_read_tags,False))

                # calculate number of changes
                num_tag_changes = len(tag_data_differences)
                
                #write data to plc
                if num_tag_changes > 0:
                    # format output based on number of changes
                    if num_tag_changes >= 2:
                        print("Writing " + str(num_tag_changes) + " tag changes to instances of " + aoi)
                    else:
                        print("Writing " + str(num_tag_changes) + " tag change to instances of " + aoi) 
                        
                    # store results of PLC write
                    write_result = write_plc_row(plc,tag_data_differences)  
                    
                    # add to failed tags list if we can't find the tag
                    if not all(write_result):

                        # extract tag name from list of tuples
                        tag_difference_list = [t[0] for t in tag_data_differences]

                        failed_write_tags = get_failed_tags(tag_difference_list,write_result)

                        # print to command line if we couldn't write any tags
                        print(failed_tag_formatter(failed_write_tags,True))

                else:
                        print("No differences for instances of " + aoi)
            else:
                print("No instances of " + aoi + " in " + plc_name + " PLC.")

        print("Finished writing to " + plc_name + " PLC.")

    plc.close()
    book.close()

if __name__ == "__main__":
    main()